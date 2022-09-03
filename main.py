from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder

in_filename = "alle-poster-2022-09-03.csv"
out_filename = "spiir-accounting-2022.xlsx"
bugfix_filename = "bugfix-file.csv"


def fix_splitting(dff: pd.DataFrame) -> pd.DataFrame:
    if len(dff.index) < 2:
        return dff
    else:
        if (
            dff.iloc[0]["MainCategoryName"] != "Hide"
            or dff.iloc[0]["CategoryName"] != "Exclude"
        ):
            remove_id = dff.iloc[0]["Id"]
            with Path(bugfix_filename).open("a") as f:
                f.write(f"{remove_id}\n")
        return dff


def format_spiir_sheet(filename: str) -> None:
    wb = openpyxl.load_workbook(filename)
    ws = wb["Sheet1"]
    max_row = ws.max_row

    my_format = "# ##0;-# ##0;0;@"
    for row in ws.iter_rows(min_row=2, max_row=70, min_col=2, max_col=14):
        for cell in row:
            cell.number_format = my_format

    dim_holder = DimensionHolder(worksheet=ws)
    dim_holder["A"] = ColumnDimension(ws, min=1, max=1, width=30)
    for col in range(ws.min_column + 1, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(
            ws, min=col, max=col, width=9
        )
    ws.column_dimensions = dim_holder

    ws.cell(row=max_row + 1, column=2).value = f"=SUM(B2:B{max_row})"
    wb.save("formatted.xlsx")


def main():
    df = pd.read_csv(
        in_filename,
        encoding="latin_1",
        sep=";",
        decimal=",",
        parse_dates=["Date", "CustomDate"],
        dayfirst=True,
        true_values=["Yes"],
        false_values=["No"],
        usecols=[
            "Id",
            "Date",
            "Description",
            "MainCategoryName",
            "CategoryName",
            "CategoryType",
            "ExpenseType",
            "Amount",
            "Extraordinary",
            "SplitGroupId",
            "CustomDate",
        ],
        dtype={
            "Id": "string",
            "Description": "string",
            "MainCategoryName": "string",
            "CategoryName": "string",
            "CategoryType": "string",
            "ExpenseType": "string",
            "SplitGroupId": "string",
        },
    )

    # Calculate CorrectedDate
    df["CorrectedDate"] = df["CustomDate"].where(df["CustomDate"].notnull(), df["Date"])

    # Delete old bugfix file
    Path(bugfix_filename).unlink(missing_ok=True)

    # Identify rows that have to be fixed
    df.groupby("SplitGroupId", as_index=False).apply(fix_splitting)

    # Identify
    df2 = df.set_index("Id")
    bugfix_file = Path(bugfix_filename)
    if bugfix_file.is_file():
        remove_ids = bugfix_file.read_text().splitlines()
        df2.drop(remove_ids, inplace=True)

    df2.reset_index(drop=True, inplace=True)
    print(f"Shape after fixing: {df2.shape}")

    df2 = df2[(df2["CategoryType"] != "Exclude") & (df2["Extraordinary"] is False)]

    df2_2021 = df2[df2["CorrectedDate"].dt.year == 2022]

    pivot2 = pd.pivot_table(
        df2_2021,
        values="Amount",
        index="CategoryName",
        columns=pd.Grouper(key="CorrectedDate", freq="M"),
        aggfunc=np.sum,
        fill_value=0,
    )
    pivot2.columns = pd.to_datetime(pivot2.columns).strftime("%b %Y")

    pivot2.to_excel(out_filename)
    print("Finished writing spreadsheet.")


if __name__ == "__main__":
    main()
    format_spiir_sheet(out_filename)
