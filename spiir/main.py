from decimal import Decimal
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.dimensions import DimensionHolder

year = 2024
in_filename = "alle-poster-2024-12-26.csv"
out_filename = f"spiir-accounting-{year}.xlsx"


def read_transactions_file(filepath: Path) -> pd.DataFrame:
    return pd.read_csv(
        filepath,
        index_col=0,
        sep=";",
        decimal=",",
        parse_dates=["Date", "CustomDate"],
        dayfirst=True,
        true_values=["Yes"],
        false_values=["No"],
        usecols=[
            "Id",
            "Date",
            "Description",
            "MainCategoryName",
            "CategoryName",
            "CategoryType",
            "ExpenseType",
            "Amount",
            "Extraordinary",
            "SplitGroupId",
            "CustomDate",
        ],
        dtype={
            "Id": "string",
            "Description": "string",
            "MainCategoryName": "string",
            "CategoryName": "string",
            "CategoryType": "string",
            "ExpenseType": "string",
            "SplitGroupId": "string",
        },
    )


def fix_split_transactions(df: pd.DataFrame) -> pd.DataFrame:
    """Fix errors in split transactions.

    There is a bug in Spiir which sometimes makes the original transaction in a
    split transaction visible. It should be hidden. Solve this be removing the
    original transaction (the first one) in each split group. Then add the non-split
    transactions.

    Args:
        df: DataFrame with raw Spiir transaction data. Expected to have a column called
            "SplitGroupId" which indicates split groups. Transactions with null values
            in "SplitGroupId" are considered as non-split transactions.

    Returns:
        A DatFrame with the fixed list of transactions.
    """
    split_group_df = df.groupby("SplitGroupId", as_index=False, group_keys=False).apply(
        lambda group: group.iloc[1:], include_groups=False
    )
    no_split_group_df = df[df.SplitGroupId.isnull()]
    return pd.concat([split_group_df, no_split_group_df])


def remove_excluded_and_extraordinary(df: pd.DataFrame) -> pd.DataFrame:
    """Filter out excluded and extraordinary rows in the dataframe.

    Args:
        df: The dataframe that contains a "CategoryType" column to specify
            category types and an "Extraordinary" column to indicate extraordinary entries.

    Returns:
        pd.DataFrame: A filtered dataframe excluding the "Exclude" and "Extraordinary" rows.
    """
    return df[
        (df["CategoryType"] != "Exclude") & (df["Extraordinary"] == False)  # noqa: E712
    ]


def correct_dates_by_year(df: pd.DataFrame, year: int) -> pd.DataFrame:
    df = df.copy()  # Ensure the input DataFrame is not a slice
    df["CorrectedDate"] = df["CustomDate"].combine_first(df["Date"])
    return df[df["CorrectedDate"].dt.year == year]


def monthly_totals(df: pd.DataFrame) -> pd.DataFrame:
    df["Amount"] = df["Amount"].apply(
        Decimal
    )  # Convert to Decimal to avoid floating point precision problems

    category_table = pd.pivot_table(
        df,
        values="Amount",
        index="CategoryName",
        columns=pd.Grouper(key="CorrectedDate", freq="ME"),
        aggfunc="sum",
        fill_value=0,
    )
    category_table.columns = pd.to_datetime(category_table.columns).strftime("%b %Y")

    # Convert back to float after calculations are done
    for col in category_table.columns:
        category_table[col] = category_table[col].astype(float)
    return category_table


def format_spiir_sheet(filename: str) -> None:
    wb = openpyxl.load_workbook(filename)
    ws = wb["Sheet1"]
    max_row = ws.max_row
    max_col = ws.max_column

    # Add row sums
    row_sum_header = ws.cell(row=max_row + 1, column=1, value="Sum")
    row_sum_header.font = Font(bold=True)
    row_sum_header.alignment = Alignment(horizontal="center")
    for col in range(ws.min_column + 1, max_col + 1):
        col_letter = get_column_letter(col)
        sum_formula = f"=SUM({col_letter}2:{col_letter}{max_row})"
        ws.cell(row=max_row + 1, column=col).value = sum_formula

    # Add col sums
    col_sum_header = ws.cell(row=1, column=max_col + 1, value="Sum")
    col_sum_header.font = Font(bold=True)
    col_sum_header.alignment = Alignment(horizontal="center")
    for row in range(ws.min_row + 1, max_row + 2):
        col_letter = get_column_letter(max_col)
        sum_formula = f"=SUM(B{row}:{col_letter}{row})"
        ws.cell(row=row, column=max_col + 1).value = sum_formula

    my_format = "# ##0;-# ##0;0;@"
    for row in ws.iter_rows(
        min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column
    ):
        for cell in row:
            cell.number_format = my_format

    dim_holder = DimensionHolder(worksheet=ws)
    dim_holder["A"] = ColumnDimension(ws, min=1, max=1, width=32)
    for col in range(ws.min_column + 1, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(
            ws, min=col, max=col, width=9
        )
    ws.column_dimensions = dim_holder

    wb.save(f"formatted-{year}.xlsx")


def main(filepath: Path = Path(__file__).parent / in_filename) -> pd.DataFrame:
    transactions_df = read_transactions_file(filepath)
    df_corrected = fix_split_transactions(transactions_df)
    df_base = remove_excluded_and_extraordinary(df_corrected)
    df_year = correct_dates_by_year(df_base, year)
    print(f"Shape after fixing: {df_year.shape}")

    category_table = monthly_totals(df_year)
    category_table.to_excel(out_filename)
    format_spiir_sheet(out_filename)
    print("Finished writing spreadsheet.")
    # category_table.to_parquet(Path(__file__).parent / "month_facit.parquet")
    return category_table


if __name__ == "__main__":
    main()
